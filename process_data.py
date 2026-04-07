"""
process_data.py  —  Allsvenskan Analytics datapipeline
Kör med: python process_data.py
Kräver:  pip install openpyxl

Hanterar:
  - Standard Stats, Shooting, Goalkeeping, Miscellaneous, Playing Time
  - xG / Expected Goals (om filen finns)
  - xA / xP (om de finns)
  - Svenska Cupen (om mappen cup/ eller Svenska Cupen/ finns)
  - 2001, 2023, 2024, 2025, 2026
"""

import os, json, datetime, math, glob
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
OUT  = os.path.join(BASE, "data.json")

# ── Typkonvertering ────────────────────────────────────────────────────────────
def to_float(v):
    if v is None or v == '' or v == 'Matches': return 0.0
    if isinstance(v, datetime.time):
        return round(v.hour + v.minute / 100.0 + v.second / 10000.0, 4)
    if isinstance(v, str):
        s = v.strip().replace(',', '.')
        if not s or s in ('-','—','N/A'): return 0.0
        try:   return float(s)
        except: return 0.0
    try: return float(v)
    except: return 0.0

def to_int(v): return int(to_float(v))

def fix_min(v):
    f = to_float(v)
    # FBRef stores minutes in thousands with Swedish decimal comma (e.g., 2.675 = 2675 min)
    # BUT only if the value is non-integer — e.g. 6 minutes is stored as integer 6 (not 0.006)
    if 0 < f < 200 and f != int(f):
        return int(round(f * 1000))
    return int(f)

def clean_nation(v):
    if not v: return ''
    parts = str(v).strip().split()
    return parts[-1].upper() if parts else ''

def clean_squad(v):
    if not v: return ''
    s = str(v).strip()
    # Normalize common variations
    norm = {'Göteborg':'IFK Göteborg','IFK Göteborg':'IFK Göteborg'}
    return norm.get(s, s)

# ── Läs xlsx ───────────────────────────────────────────────────────────────────
def read_sheet(path):
    if not path or not os.path.exists(path): return []
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        print(f"    ⚠ Kunde inte läsa {os.path.basename(path)}: {e}")
        return []
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return []
    # Find header row (contains 'Player' or 'Squad')
    header_idx = 0
    for i, row in enumerate(rows[:4]):
        if 'Player' in row or 'Squad' in row or 'Rk' in row:
            header_idx = i; break
    header = rows[header_idx]
    # Deduplicate column names
    seen, cols = {}, []
    for c in header:
        key = str(c) if c else '_'
        cnt = seen.get(key, 0) + 1
        seen[key] = cnt
        cols.append(f"{key}_{cnt}" if cnt > 1 else key)
    records = []
    main_header_set = set(c for c in cols if c and c != '_')
    for row in rows[header_idx + 1:]:
        if not any(row): continue
        d = dict(zip(cols, row))
        player_val = d.get("Player")
        squad_val  = d.get("Squad")
        # Detect mid-table header row with DIFFERENT columns (different table section)
        # e.g. FBRef appends Playing Time table after Shooting table in same file
        if player_val == "Player" and squad_val == "Squad":
            # Check if column structure differs from main header
            new_cols = [str(v).strip() if v else '' for v in row[:len(cols)]]
            new_set  = set(c for c in new_cols if c and c not in ('Rk','Player','Nation','Pos','Squad','Age','Born','_'))
            old_set  = set(c for c in cols if c not in ('Rk','Player','Nation','Pos','Squad','Age','Born','_'))
            if new_set != old_set and len(new_set & old_set) < len(old_set) // 2:
                break  # Different table structure — stop reading
            continue  # Same structure repeat header — skip and continue
        if player_val in ('Player', 'Matches'): continue
        if squad_val  in ('Squad',):            continue
        if player_val is None and squad_val is None: continue
        records.append(d)
    return records

import unicodedata as _ud

def _norm(name):
    """Normalize name: decompose unicode, replace special chars, lowercase."""
    if not name: return ""
    s = _ud.normalize("NFKD", str(name).strip())
    s = "".join(c for c in s if not _ud.combining(c))
    for old_c, new_c in [("Þ","th"),("þ","th"),("ð","d"),("Ð","D"),
                          ("ø","o"),("Ø","O"),("æ","ae"),("Æ","AE"),
                          ("ß","ss"),("ł","l"),("Ł","L")]:
        s = s.replace(old_c, new_c)
    return s.lower().strip()

def idx(records, key="Player"):
    """Build multi-key name index for robust fuzzy matching."""
    out = {}
    for r in records:
        raw = r.get(key, "")
        if not raw: continue
        out[raw] = r                          # exact
        norm = _norm(raw)
        out[norm] = r                          # normalized
        parts = norm.split()
        if parts:
            out[parts[-1]] = r                 # last name
        if len(parts) >= 2:
            out[f"{parts[0]} {parts[-1]}"] = r # first + last
    return out

def _lookup(name, index):
    """Try exact → normalized → first+last → last-name to find player."""
    if not name: return None
    if name in index: return index[name]
    norm = _norm(name)
    if norm in index: return index[norm]
    parts = norm.split()
    if len(parts) >= 2 and f"{parts[0]} {parts[-1]}" in index:
        return index[f"{parts[0]} {parts[-1]}"]
    if parts and parts[-1] in index:
        return index[parts[-1]]
    return None
# ── xG/xA/xP fil-parser (abbreviated names: "M. Fenger" format) ──────────────

def _dedupe_name(raw):
    """Hantera dubblade namn: 'M. FengerM. Fenger' → 'M. Fenger'"""
    if not raw: return ""
    s = str(raw).strip()
    n = len(s)
    if n < 3: return s
    def _n(t):
        t = _ud.normalize('NFKD', t)
        t = ''.join(c for c in t if not _ud.combining(c))
        return t.lower().strip()
    for split in range(max(2, n//4), min(n-1, 3*n//4 + 1)):
        first, second = s[:split], s[split:]
        nf, ns = _n(first), _n(second)
        if nf == ns:
            return first.strip()
        # Truncation case: second is cut-off version of first
        if len(nf) >= 4 and ns.startswith(nf[:len(nf)//2+2]) and first[0] == second[0]:
            return first.strip()
    return s

def _abbrev_key(name):
    """'M. Fenger' → ('m','fenger') matchningsnyckel"""
    s = _dedupe_name(name)
    s = _ud.normalize('NFKD', s)
    s = ''.join(c for c in s if not _ud.combining(c))
    for o2, n2 in [('ć','c'),('Ć','C'),('š','s'),('Š','S'),
                   ('ž','z'),('č','c'),('ø','o'),('æ','ae')]:
        s = s.replace(o2, n2)
    s = s.strip().lower()
    parts = s.split()
    if not parts: return None
    if len(parts) >= 2:
        return (parts[0].rstrip('.')[0], parts[-1])
    return (parts[0][0], parts[0])

def _fullname_key(name):
    """'Max Fenger' → ('m','fenger') matchningsnyckel"""
    s = _norm(name)  # uses existing _norm from idx()
    parts = s.split()
    if not parts: return None
    if len(parts) >= 2:
        return (parts[0][0], parts[-1])
    return (parts[0][0], parts[0])

def _strip_q(v):
    """Strip surrounding quotes from Excel string values."""
    if v is None: return None
    s = str(v).strip()
    if len(s) >= 2 and s[0] == "'" and s[-1] == "'": return s[1:-1]
    return s

def read_xg_file(path):
    """Läs xG/xA/xP-fil med förkortade, dubblade namn.
    Hanterar både col0 och col1 som namnkolumn, samt citattecken.
    Returnerar dict: (initial, efternamn) → {xG, xA, xP, xPpx}
    """
    if not path or not os.path.exists(path): return {}
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        print(f"    ⚠ Kunde inte läsa {os.path.basename(path)}: {e}")
        return {}
    ws  = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return {}

    # Detect which column has player names (col 0 or col 1)
    # and where xG/xA/xP columns are
    name_col = 1  # default (old format)
    xg_col, xa_col, xp_col, xp90_col = 5, 6, 7, 8

    # Check header row
    for i, row in enumerate(rows[:3]):
        cols = [_strip_q(v) for v in row]
        if 'SPELARE' in str(cols[0]):
            name_col = 0  # new format: name in col 0
        for j, v in enumerate(cols):
            sv = str(v or '').upper()
            if sv == 'XG':   xg_col  = j
            elif sv == 'XA': xa_col  = j
            elif sv == 'XP': xp_col  = j
            elif 'XP/90' in sv or sv == 'XP90': xp90_col = j

    result = {}
    for row in rows:
        raw_name = _strip_q(row[name_col]) if len(row) > name_col else None
        if not raw_name or raw_name.upper() in ('SPELARE', '', 'PLAYER'): continue

        name = _dedupe_name(raw_name)
        key  = _abbrev_key(name)
        if not key: continue
        try:
            result[key] = {
                "xG":   to_float(row[xg_col]   if len(row) > xg_col   else 0),
                "xA":   to_float(row[xa_col]   if len(row) > xa_col   else 0),
                "xP":   to_float(row[xp_col]   if len(row) > xp_col   else 0),
                "xPpx": to_float(row[xp90_col] if len(row) > xp90_col else 0),
                "_xg_name": name,
            }
        except Exception:
            continue
    return result

def lookup_xg(full_name, xg_idx):
    """Matcha fullständigt namn mot xG-indexet."""
    if not xg_idx: return None
    key = _fullname_key(full_name)
    if not key: return None
    return xg_idx.get(key)



# ── Hitta fil med glob (hanterar namnvariationer) ──────────────────────────────
def find_file(base_dir, patterns):
    for pat in patterns:
        hits = glob.glob(os.path.join(base_dir, pat))
        if hits: return hits[0]
    return None

# ── Bygg spelarobjekt ──────────────────────────────────────────────────────────
def build_player(std, sh, misc, gk, pt, xg_row, competition="allsvenskan"):
    p = {}
    src = std or sh or misc or gk or pt or {}
    p["name"]        = str(src.get("Player","")).strip()
    p["nation"]      = clean_nation(src.get("Nation",""))
    p["pos"]         = str(src.get("Pos","")).strip()
    p["squad"]       = clean_squad(src.get("Squad",""))
    raw_age = src.get("Age", 0)
    if raw_age and isinstance(raw_age, str) and "-" in str(raw_age):
        raw_age = str(raw_age).split("-")[0]  # "27-280" → "27"
    p["age"]         = to_int(raw_age)
    p["competition"] = competition

    if std:
        p["mp"]       = to_int(std.get("MP"))
        p["starts"]   = to_int(std.get("Starts"))
        p["min"]      = fix_min(std.get("Min"))
        p["nineties"] = to_float(std.get("90s"))
        p["gls"]      = to_int(std.get("Gls"))
        p["ast"]      = to_int(std.get("Ast"))
        p["gPluA"]    = to_int(std.get("G+A"))
        p["gMinusPK"] = to_int(std.get("G-PK"))
        p["pk"]       = to_int(std.get("PK"))
        p["pkAtt"]    = to_int(std.get("PKatt"))
        p["crdY"]     = to_int(std.get("CrdY"))
        p["crdR"]     = to_int(std.get("CrdR"))
        p["glsPer90"] = to_float(std.get("Gls_2"))
        p["astPer90"] = to_float(std.get("Ast_2"))

    if sh:
        p["sh"]       = to_int(sh.get("Sh"))
        p["sot"]      = to_int(sh.get("SoT"))
        p["sotPct"]   = to_float(sh.get("SoT%"))
        p["shPer90"]  = to_float(sh.get("Sh/90"))
        p["sotPer90"] = to_float(sh.get("SoT/90"))
        p["gPerSh"]   = to_float(sh.get("G/Sh"))
        p["gPerSoT"]  = to_float(sh.get("G/SoT"))
        if "nineties" not in p: p["nineties"] = to_float(sh.get("90s"))
        if "gls" not in p:      p["gls"]      = to_int(sh.get("Gls"))

    if misc:
        p["fls"]    = to_int(misc.get("Fls"))
        p["fld"]    = to_int(misc.get("Fld"))
        p["off"]    = to_int(misc.get("Off"))
        p["crs"]    = to_int(misc.get("Crs"))
        p["int"]    = to_int(misc.get("Int"))
        p["tklW"]   = to_int(misc.get("TklW"))
        p["pkWon"]  = to_int(misc.get("PKwon"))
        p["pkCon"]  = to_int(misc.get("PKcon"))
        p["og"]     = to_int(misc.get("OG"))
        if "crdY"     not in p: p["crdY"] = to_int(misc.get("CrdY"))
        if "nineties" not in p: p["nineties"] = to_float(misc.get("90s"))

    if gk:
        p["gkGA"]      = to_int(gk.get("GA"))
        p["gkGA90"]    = to_float(gk.get("GA90"))
        p["gkSoTA"]    = to_int(gk.get("SoTA"))
        p["gkSaves"]   = to_int(gk.get("Saves"))
        p["gkSavePct"] = to_float(gk.get("Save%"))
        p["gkW"]       = to_int(gk.get("W"))
        p["gkD"]       = to_int(gk.get("D"))
        p["gkL"]       = to_int(gk.get("L"))
        p["gkCS"]      = to_int(gk.get("CS"))
        p["gkCSPct"]   = to_float(gk.get("CS%"))

    if pt:
        if "mp"       not in p: p["mp"]      = to_int(pt.get("MP"))
        if "starts"   not in p: p["starts"]  = to_int(pt.get("Starts"))
        if "min"      not in p: p["min"]     = fix_min(pt.get("Min"))
        if "nineties" not in p: p["nineties"]= to_float(pt.get("90s"))

    # xG / xA / xP
    if xg_row:
        p["xG"]   = to_float(xg_row.get("xG")   or xg_row.get("npxG")  or 0)
        p["xA"]   = to_float(xg_row.get("xAG")  or xg_row.get("xA")    or 0)
        p["npxG"] = to_float(xg_row.get("npxG") or xg_row.get("xG")    or 0)
        # per-90 versions (FBRef stores them in column 2 with same name)
        p["xGpx"] = to_float(xg_row.get("xG_2")  or xg_row.get("npxG_2") or 0)
        p["xApx"] = to_float(xg_row.get("xAG_2") or xg_row.get("xA_2")   or 0)
        # xP = expected pass completion (FBRef passing advanced)
        # Field names vary: xP, xPass, passer_xp, Cmp_2 in some exports
        p["xP"]   = to_float(xg_row.get("xP")    or xg_row.get("xPass") or
                               xg_row.get("passer_xp") or 0)
        # Some FBRef exports have xP as a per-90 value directly
        p["xPpx"] = to_float(xg_row.get("xP_2") or xg_row.get("xPass_2") or 0)
        # Goal vs xG difference
        xg = p.get("xG",0) or p.get("npxG",0)
        if xg and p.get("gls"):
            p["gMinusXG"] = round(p["gls"] - xg, 2)
        # xA vs actual assists diff
        xa = p.get("xA",0)
        if xa and p.get("ast"):
            p["aMinusXA"] = round(p["ast"] - xa, 2)

    # Per-90 computed fields
    nineties = max(p.get("nineties",0), 0.1)
    for raw_key, per_key in [
        ("fls","flsPer90"), ("fld","fldPer90"),
        ("int","intPer90"), ("tklW","tklWPer90"),
        ("crs","crsPer90"), ("xA","xApx"),  ("xP","xPpx"),
    ]:
        if raw_key in p and per_key not in p:
            p[per_key] = round(p[raw_key] / nineties, 3)
    if "xG" in p and "xGpx" not in p:
        p["xGpx"] = round(p["xG"] / nineties, 3)

    return p

# ── Percentiler ────────────────────────────────────────────────────────────────
PERCENTILE_METRICS = {
    "GK": ["gkSavePct","gkGA90","gkCSPct","gkSaves","gkW","min","gkSoTA"],
    "DF": ["min","intPer90","tklWPer90","flsPer90","crsPer90","glsPer90","astPer90","crdY","xGpx","xApx"],
    "MF": ["glsPer90","astPer90","shPer90","sotPct","intPer90","tklWPer90","fldPer90","crdY","xGpx","xApx","xPpx"],
    "FW": ["glsPer90","astPer90","shPer90","sotPct","gPerSh","gPerSoT","fldPer90","off","xGpx","xApx","xPpx","gMinusXG"],
}
INVERT = {"gkGA90","crdY","crdR","flsPer90","off","pkCon","gkSoTA"}

def pct_rank(val, vals, invert=False):
    nz = [v for v in vals if isinstance(v,(int,float)) and not math.isnan(v) and v > 0]
    if not nz: return 0
    p = round(sum(1 for v in nz if v <= val) / len(nz) * 100)
    return max(0, min(100, 100 - p if invert else p))

def pos_group(pos):
    if not pos: return "U"
    p = str(pos).split(",")[0].strip()
    return {"GK":"GK","DF":"DF","MF":"MF","FW":"FW"}.get(p,"U")

def add_percentiles(players):
    groups = {"GK":[],"DF":[],"MF":[],"FW":[],"U":[]}
    for p in players:
        g = pos_group(p.get("pos",""))
        groups.setdefault(g, []).append(p)

    all_metrics = set()
    for v in PERCENTILE_METRICS.values(): all_metrics |= set(v)

    for group, gps in groups.items():
        if not gps: continue
        mets = PERCENTILE_METRICS.get(group, list(all_metrics))
        for metric in all_metrics:
            vals = [p.get(metric,0) for p in gps]
            inv  = metric in INVERT
            for p in gps:
                p.setdefault("pct",{})[metric] = pct_rank(p.get(metric,0), vals, inv)

    return players

# ── Liga-snitt per position ───────────────────────────────────────────────────
def compute_league_averages(players):
    groups = {"GK":[],"DF":[],"MF":[],"FW":[],"ALL":[]}
    for p in players:
        g = pos_group(p.get("pos",""))
        groups.setdefault(g,[]).append(p)
        groups["ALL"].append(p)

    avgs = {}
    num_metrics = [
        "gls","ast","gPluA","mp","min","nineties","sh","sot","sotPct","shPer90","sotPer90",
        "gPerSh","gPerSoT","int","intPer90","tklW","tklWPer90","fls","flsPer90",
        "fld","fldPer90","crs","crsPer90","off","crdY","crdR","glsPer90","astPer90",
        "xG","xA","xGpx","xApx","npxG","gMinusXG","xP","xPpx","aMinusXA",
        "gkSavePct","gkGA90","gkCS","gkCSPct","gkSaves","gkSoTA","gkW","gkD","gkL",
    ]
    for group, gps in groups.items():
        if not gps: continue
        avgs[group] = {}
        for m in num_metrics:
            vals = [p.get(m,0) for p in gps if p.get(m) and p.get(m) != 0]
            avgs[group][m] = round(sum(vals)/len(vals), 3) if vals else 0
    return avgs

# ── Squad ──────────────────────────────────────────────────────────────────────
def build_squad(row, sh_row=None, xg_row=None):
    s = {
        "squad":   clean_squad(row.get("Squad","")),
        "numPl":   to_int(row.get("# Pl")),
        "age":     to_float(row.get("Age")),
        "poss":    to_float(row.get("Poss")),
        "mp":      to_int(row.get("MP")),
        "gls":     to_int(row.get("Gls")),
        "ast":     to_int(row.get("Ast")),
        "gPluA":   to_int(row.get("G+A")),
        "crdY":    to_int(row.get("CrdY")),
        "crdR":    to_int(row.get("CrdR")),
    }
    if sh_row:
        s["sh"]     = to_int(sh_row.get("Sh"))
        s["sot"]    = to_int(sh_row.get("SoT"))
        s["sotPct"] = to_float(sh_row.get("SoT%"))
    if xg_row:
        s["xG"]  = to_float(xg_row.get("xG") or xg_row.get("npxG") or 0)
        s["xA"]  = to_float(xg_row.get("xAG") or xg_row.get("xA") or 0)
    return s

# ── Säsongskonfiguration ───────────────────────────────────────────────────────
def auto_find(year_dir, patterns):
    for pat in patterns:
        hits = glob.glob(os.path.join(year_dir, pat))
        if hits: return hits[0]
    return None

def process_competition(comp_dir, year, competition="allsvenskan"):
    """Process one competition folder."""
    if not os.path.isdir(comp_dir): return [], []

    print(f"      → {competition} ({os.path.basename(comp_dir)})")

    def af(pats): return auto_find(comp_dir, pats)

    std_r  = idx(read_sheet(af([f"*Standard Stats*{year}*", f"*Standard*{year}*", "*Standard Stats*","*standard*"])))
    sh_r   = idx(read_sheet(af([f"*Shooting*{year}*", "*Shooting*","*shooting*"])))
    misc_r = idx(read_sheet(af([f"*Miscellaneous*{year}*", "*Miscellaneous*","*misc*","*Misc*"])))
    gk_r   = idx(read_sheet(af([f"*Goalkeeping*{year}*", f"*Goalkeeper*{year}*", "*Goalkeeping*","*goalkeeping*","*GK*"])))
    pt_r   = idx(read_sheet(af([f"*Playing Time*{year}*", "*Playing Time*","*playing*"])))
    xg_r   = idx(read_sheet(af([
        "*Expected*","*xG*","*xg*","*Advanced*","*advanced*"
    ])))
    # xG/xA/xP-fil med förkortade namn (t.ex. "xG_ xA _ xP 2025.xlsx")
    xgap_path = auto_find(comp_dir, ["xG, xA*","xG*.xlsx","*xG*xA*","*xG*xP*","*xA*xP*","xG_*","*xG_*","*xG*.xlsx"])
    xgap_idx  = read_xg_file(xgap_path) if xgap_path else {}

    names = set(std_r)|set(sh_r)|set(misc_r)|set(gk_r)|set(pt_r)
    players = []
    for name in names:
        pl = build_player(
            _lookup(name, std_r),
            _lookup(name, sh_r),
            _lookup(name, misc_r),
            _lookup(name, gk_r),
            _lookup(name, pt_r),
            _lookup(name, xg_r),
            competition=competition,
        )
        # Komplettera med xG/xA/xP från förkortad namnfil om det saknas
        if pl and pl.get("name") and not pl.get("xG",0) and xgap_idx:
            xg_d = lookup_xg(pl["name"], xgap_idx)
            if xg_d:
                n90 = max(pl.get("nineties",0), 0.1)
                pl["xG"]   = xg_d["xG"]
                pl["xA"]   = xg_d["xA"]
                pl["xP"]   = xg_d["xP"]
                pl["xPpx"] = xg_d["xPpx"]
                pl["xGpx"] = round(xg_d["xG"] / n90, 3)
                pl["xApx"] = round(xg_d["xA"] / n90, 3)
                if xg_d["xG"] and pl.get("gls"):
                    pl["gMinusXG"] = round(pl["gls"] - xg_d["xG"], 2)
        if pl.get("name") and pl.get("squad"):
            players.append(pl)

    # Squad
    sq_std = auto_find(comp_dir, ["*Squad*Standard*","*squad*standard*","Squad Standard*"])
    sq_sh  = auto_find(comp_dir, ["*Squad*Shoot*","*squad*shoot*","Squad Shoot*"])
    sq_xg  = auto_find(comp_dir, ["*Squad*Expected*","*Squad*xG*"])
    sq_rows = read_sheet(sq_std) if sq_std else []
    sh_idx  = {r.get("Squad"):r for r in (read_sheet(sq_sh)  or []) if r.get("Squad")}
    xg_idx  = {r.get("Squad"):r for r in (read_sheet(sq_xg)  or []) if r.get("Squad")}
    squads  = []
    for row in sq_rows:
        if not row.get("Squad"): continue
        sq = build_squad(row, sh_idx.get(row["Squad"]), xg_idx.get(row["Squad"]))
        if sq["squad"]: squads.append(sq)

    return players, squads

YEAR_DIRS = ["2001","2022","2023","2024","2025","2026"]

# ── Extra statistik (passningar, löpningar etc.) ──────────────────────────────

def _dedupe_extra(raw):
    """Same deduplication as _dedupe_name but standalone."""
    if not raw: return ""
    s = str(raw).strip()
    n = len(s)
    if n < 3: return s
    for split in range(max(2,n//4), min(n-1,3*n//4+1)):
        first, second = s[:split], s[split:]
        if _norm(first) == _norm(second): return first.strip()
        if len(_norm(first))>=4 and _norm(second).startswith(_norm(first)[:len(_norm(first))//2+2]) and first[0]==second[0]:
            return first.strip()
    return s

def _extra_abbrev_key(name):
    s = _norm(_dedupe_extra(name))
    parts = s.strip().split()
    if not parts: return None
    return (parts[0].rstrip(".")[0], parts[-1])

def safe_f(v):
    if v is None or (isinstance(v, str) and v.strip() == ""): return None
    try: return float(str(v).replace(",","."))
    except: return None

def parse_extra_file(path):
    """Parse Passning_YYYY.xlsx or similar multi-sheet file.
    Returns: (player_stats_dict, domare_list)
    """
    if not path or not os.path.exists(path): return {}, []
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        print(f"    ⚠ Extra-fil fel: {e}"); return {}, []

    # Stöder både 2025-format (Passning_2025) och 2026-format (Statistik_2026)
    SHEET_COLS = {
        # ── 2025-format ──────────────────────────────
        "PASSNINGAR & PASSNINGSPROCENT": {"pass_total":5,"pass_complete":6,"pass_pct":7},
        "Framåtpassningar":              {"fwd_pass":5,"fwd_pass_comp":6,"fwd_pass_pct":7},
        "Huvudspel & luftdueller":       {"headers_won":5,"headers_pct":6,"np_duels":7,"np_duels_pct":8},
        "Progressiva passningar":        {"prog_pass":5,"prog_pass_pct":6,"prog_pass_p90":7},
        "ACC. & PROGRESSIVA LÖPNINGAR  i":{"acc_carries":5,"prog_carries":6,"acc_p90":7,"prog_carries_p90":8},
        "Återerövringar":                {"recoveries":5,"recoveries_p90":6},
        "Lång passningar & genomskärare":{"long_pass":5,"through_balls":6,"long_through_p90":7},
        # ── 2026-format ──────────────────────────────
        "MÅL & xG":                      {"gls_file":5,"shot_acc_pct":6,"glsPer90_file":7},
        "xG, xA & xP":                   {"xG":5,"xA":6,"xP":7},
        "ASSIST & xA":                    {"ast_file":5,"ast_pct":6,"astPer90_file":7},
        "Poäng & xP":                     {"g_plus_a_file":5,"xP_file":6},
        "Hockeyassist":                   {"second_ast":5,"second_ast_p90":6},
        "Målchanser":                     {"chances_created":5,"chances_p90":6},
        "Nyckelpassningar":               {"key_passes":6,"key_passes_p90":7},
        "PROGRESSIVA PASSNINGAR":         {"prog_pass":5,"prog_pass_pct":6,"prog_pass_p90":7},
        "Långa passningar & genomskärare":{"long_pass":5,"through_balls":6,"long_through_p90":7},
        "Duellerspel":                    {"od_duels":5,"od_duels_pct":6,"dd_duels":7},
        "Återerövringar":                 {"recoveries":5,"recoveries_p90":6},
        "Offsides":                       {"off_file":5,"off_p90":6},
        "Kort":                           {"crdY_file":5,"crdR_file":6,"crdY_p90":7},
        "Acc. och Progressiva löpningar": {"acc_carries":5,"prog_carries":6,"acc_p90":7,"prog_carries_p90":8},
    }
    player_data = {}
    domare_list = []

    for sname, col_map in SHEET_COLS.items():
        if sname not in wb.sheetnames: continue
        ws = wb[sname]
        for row in list(ws.iter_rows(values_only=True))[1:]:
            raw = row[1] if len(row) > 1 else row[0]
            if not raw: continue
            name = _dedupe_extra(str(raw))
            key  = _extra_abbrev_key(name)
            if not key: continue
            if key not in player_data: player_data[key] = {"_abbrev": name}
            for stat_key, col_idx in col_map.items():
                v = safe_f(row[col_idx]) if col_idx < len(row) else None
                if v is not None: player_data[key][stat_key] = v

    # Domare sheet
    if "Domare" in wb.sheetnames:
        ws = wb["Domare"]
        for row in list(ws.iter_rows(values_only=True))[1:]:
            raw = row[1] if len(row) > 1 else None
            if not raw: continue
            name = _dedupe_extra(str(raw))
            if not name: continue
            domare_list.append({
                "name":      name,
                "matches":   safe_f(row[2]),
                "fouls_pm":  safe_f(row[3]),
                "yk_pm":     safe_f(row[4]),
                "rk_pm":     safe_f(row[5]),
                "pen":       safe_f(row[6]),
                "pen_pm":    safe_f(row[7]),
            })

    print(f"      → Extra stats: {len(player_data)} spelare, {len(domare_list)} domare")
    return player_data, domare_list


# ── Nationalitetsfiler ────────────────────────────────────────────────────────
def parse_nationality_file(path):
    """Parse FBRef Allsvenskan Nationalities xlsx.
    Returns list of {rank, code, nation, players, minutes, player_list}"""
    if not path or not os.path.exists(path): return []
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        print(f"    ⚠ {e}"); return []
    ws  = wb.active
    out = []
    for row in ws.iter_rows(values_only=True):
        if not row[1]: continue
        # Skip header rows (various formats FBRef uses)
        if str(row[1]).strip() in ("Nation","'Nation","# Players"): continue
        if str(row[0]).strip() in ("Rk","Rk\n▲","Rank"): continue
        try:
            nation_raw = str(row[1]).strip()
            parts      = nation_raw.split(" ", 1)
            code       = parts[0].upper().strip("'")
            name       = parts[1] if len(parts) > 1 else parts[0]
            players    = int(float(row[2])) if row[2] else 0
            # Minutes: stored in thousands if has decimals, else actual
            raw_min    = to_float(row[3])
            minutes    = int(round(raw_min * 1000)) if (raw_min > 0 and raw_min != int(raw_min)) else int(raw_min)
            player_list= str(row[4]).split(", ") if row[4] else []
            out.append({
                "code": code, "nation": name,
                "players": players, "minutes": minutes,
                "player_list": player_list,
            })
        except Exception:
            continue
    out.sort(key=lambda x: -x["players"])
    return out


# ── Main ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("Allsvenskan Analytics — Dataprocessor v2")
    output = {"seasons": {}, "generated": str(datetime.datetime.now())[:19]}

    DATA_ROOT = os.path.join(BASE, "90minutersdata")
    if not os.path.isdir(DATA_ROOT):
        print(f"⚠  90minutersdata/ hittades inte i {BASE}")
        exit(1)

    for year in YEAR_DIRS:
        year_player = os.path.join(DATA_ROOT, "Player", year)
        year_squad  = os.path.join(DATA_ROOT, "Squad",  year)

        if not os.path.isdir(year_player) and not os.path.isdir(year_squad):
            continue

        print(f"\n  Säsong {year}…")
        all_players = []
        all_squads  = []

        # ── Allsvenskan (main)
        ps, sq = process_competition(year_player, year, "allsvenskan")
        # ── Komplettera med xG/xA/xP-fil om den finns i rooten
        # Search for xG file in multiple locations
        xgap_root = None
        for search_dir in [
            DATA_ROOT,
            os.path.join(DATA_ROOT, "Player", year),
            os.path.join(DATA_ROOT, "Squad", year),
            year_player,
        ]:
            if not os.path.isdir(search_dir): continue
            xgap_root = find_file(search_dir, [
                f"xG, xA, xP {year}.xlsx",
                f"xG, xA & xP {year}.xlsx",
                f"xG, xA, xP {year}*.xlsx",
                f"*xG*xA*xP*{year}*",
                f"*{year}*xG*xA*",
                f"xG__xA___xP_{year}.xlsx",
                f"xG_ xA _ xP {year}.xlsx",
                f"*xG*{year}*.xlsx",
            ])
            if xgap_root:
                break
        if xgap_root:
            root_xgap = read_xg_file(xgap_root)
            print(f"      → xG/xA/xP root-fil: {os.path.basename(xgap_root)} ({len(root_xgap)} spelare)")
            for p in ps:
                if p.get("xG",0) > 0: continue  # already has xG
                xg_d = lookup_xg(p.get("name",""), root_xgap)
                if xg_d:
                    n90 = max(p.get("nineties",0), 0.1)
                    p["xG"]   = xg_d["xG"]
                    p["xA"]   = xg_d["xA"]
                    p["xP"]   = xg_d["xP"]
                    p["xPpx"] = xg_d["xPpx"]
                    p["xGpx"] = round(xg_d["xG"] / n90, 3)
                    p["xApx"] = round(xg_d["xA"] / n90, 3)
                    if xg_d["xG"] and p.get("gls"):
                        p["gMinusXG"] = round(p["gls"] - xg_d["xG"], 2)
                    if xg_d["xA"] and p.get("ast"):
                        p["aMinusXA"] = round(p["ast"] - xg_d["xA"], 2)
        # Merge squad data from squad folder
        sq_ps, sq_sq = process_competition(year_squad, year, "allsvenskan")
        all_players += ps
        all_squads  += sq if sq else sq_sq

        # ── Svenska Cupen (om den finns)
        cup_dirs = [
            os.path.join(DATA_ROOT, "Player", year, "Svenska Cupen"),
            os.path.join(DATA_ROOT, "Player", year, "Cup"),
            os.path.join(DATA_ROOT, "Player", year, "cup"),
            os.path.join(DATA_ROOT, "Cup",    year),
            os.path.join(DATA_ROOT, "Svenska Cupen", year),
        ]
        for cup_dir in cup_dirs:
            if os.path.isdir(cup_dir):
                cup_ps, _ = process_competition(cup_dir, year, "cup")
                # Merge cup stats onto existing player (separate key)
                cup_idx = {p["name"]: p for p in cup_ps}
                for p in all_players:
                    if p["name"] in cup_idx:
                        p["cup"] = {k:v for k,v in cup_idx[p["name"]].items()
                                    if k not in ("name","squad","nation","pos","age","competition","pct")}
                # Add cup-only players
                existing = {p["name"] for p in all_players}
                for cp in cup_ps:
                    if cp["name"] not in existing:
                        cp["competition"] = "cup"
                        all_players.append(cp)
                print(f"        ✓ Svenska Cupen: {len(cup_ps)} spelare")
                break

        if not all_players and not all_squads:
            print(f"    (inga data)")
            continue

        # Deduplicera spelare (behåll Allsvenskan om dubblett)
        seen = {}
        for p in all_players:
            k = p["name"]
            if k not in seen or p.get("competition") == "allsvenskan":
                seen[k] = p
        all_players = list(seen.values())

        all_players = add_percentiles(all_players)
        all_players.sort(key=lambda x: (x.get("squad",""), x.get("name","")))

        league_avgs = compute_league_averages(all_players)

        print(f"    ✓ {len(all_players)} spelare, {len(all_squads)} lag")
        xg_count = sum(1 for p in all_players if p.get("xG",0) > 0)
        if xg_count: print(f"    ✓ {xg_count} spelare med xG-data")

        # Nationality files (named YYYY_Allsvenskan_Nationalities.xlsx)
        nat_path = find_file(DATA_ROOT, [
            f"{year}_Allsvenskan_Nationalities.xlsx",
            f"Nationalities {year}.xlsx",
            f"*{year}*Nationalit*.xlsx",
            f"*Nationalit*{year}*.xlsx",
        ])
        nationalities = parse_nationality_file(nat_path) if nat_path else []
        if nationalities:
            print(f"    ✓ Nationaliteter: {len(nationalities)} länder")

        # ── Extra statistik-fil (passningar, löpningar etc.)
        # Sök extrastatistikfil i flera mappar (Passning_YYYY, Statistik_YYYY etc.)
        extra_path = None
        for search_dir in [
            DATA_ROOT,
            os.path.join(DATA_ROOT, "Player", year),
            os.path.join(DATA_ROOT, "Squad",  year),
            year_player,
        ]:
            if not os.path.isdir(search_dir): continue
            extra_path = find_file(search_dir, [
                f"Passning_{year}.xlsx",
                f"Statistik_{year}.xlsx",
                f"*Passning*{year}*",
                f"*Statistik*{year}*",
                f"*{year}*Passning*",
                f"*{year}*Statistik*",
                f"Passning_{year[2:]}.xlsx",   # e.g. Passning_25.xlsx
                f"*extra*{year}*",
            ])
            if extra_path:
                break
        extra_players, domare = {}, []
        if extra_path:
            extra_players, domare = parse_extra_file(extra_path)
            # Merge extra stats into player objects
            for p in all_players:
                key = _extra_abbrev_key(p.get("name",""))
                if key and key in extra_players:
                    extras = extra_players[key]
                    for k, v in extras.items():
                        if k.startswith("_"): continue
                        # Don't overwrite FBRef stats with file stats
                        # (file stats use _file suffix for duplicates)
                        if k.endswith("_file"): continue
                        # Only set if not already present from FBRef
                        if not p.get(k) or p.get(k) == 0:
                            p[k] = v
                    # If xG came from extra file, compute per90 fields
                    n90 = max(p.get("nineties", 0), 0.1)
                    if extras.get("xG") and not p.get("xGpx"):
                        p["xGpx"] = round(extras["xG"] / n90, 3)
                    if extras.get("xA") and not p.get("xApx"):
                        p["xApx"] = round(extras["xA"] / n90, 3)
                    if extras.get("xP") and not p.get("xPpx"):
                        p["xPpx"] = round(extras["xP"] / n90, 3)
                    if extras.get("xG") and p.get("gls") is not None:
                        p["gMinusXG"] = round(p.get("gls",0) - extras["xG"], 2)

        output["seasons"][year] = {
            "players":         all_players,
            "squads":          all_squads,
            "league_averages": league_avgs,
            "nationalities":   nationalities,
            "domare":          domare,
        }

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, separators=(',',':'))

    total_p  = sum(len(s["players"]) for s in output["seasons"].values())
    total_sq = sum(len(s["squads"])  for s in output["seasons"].values())
    kb = os.path.getsize(OUT) // 1024
    print(f"\n✓  data.json: {total_p} spelare, {total_sq} lag, {kb} KB")
    print(f"   Säsonger: {', '.join(sorted(output['seasons'].keys()))}")
